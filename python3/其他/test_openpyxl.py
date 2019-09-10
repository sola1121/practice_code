from faker import Faker
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


fake = Faker("zh-CN")

# 数据生成
data = list()
for _ in range(100):
    data.append([fake.ssn(), fake.name(), fake.date()])

# print(data)

wb = Workbook()
sheet = wb.get_sheet_by_name(wb.sheetnames[0])

sheet.title = "sheet1"

# print(wb.sheetnames)

for info in data:
    sheet.append(info)

file = save_virtual_workbook(wb)
# print(file)  # bytes file content

wb.save("temp.xlsx")


# 数据读取
import io
from openpyxl import load_workbook

wb_file = load_workbook(filename="temp.xlsx", read_only=True) 

print(dir(wb_file))

print(wb_file.get_sheet_names())

print(dir(wb_file.worksheets))

sheet = wb_file.get_sheet_by_name(wb_file.sheetnames[0])

for info in sheet:
    pass
else:
    print(dir(info[0]), len(info))
    print(info[0].value)
